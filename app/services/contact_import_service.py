import csv
import hashlib
import io
import re
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import PurePath
from typing import Any, Dict, List, Optional, Tuple

from bson import ObjectId
from openpyxl import load_workbook

from app.api.response_helpers import pagination_metadata
from app.errors import ApiError, ConflictError, GoneError, NotFoundError, PayloadTooLargeError, ValidationApiError
from app.models.crm_model import LeadPriority, LeadStatus, PreferredMode
from app.repositories.contact_import_repository import (
    all_import_rows,
    claim_import_commit,
    find_import_job,
    insert_import_job,
    insert_import_rows,
    list_import_rows,
    update_import_job,
    update_import_row,
)
from app.repositories.contact_repository import find_contact_by_id, find_contact_by_normalized_phone
from app.schemas.contact_import_schema import ContactImportPreviewModel
from app.schemas.contact_schema import ContactCreateModel, ContactPatchModel
from app.services.assignment_service import validate_counsellor
from app.services.audit_service import write_audit_event
from app.services.contact_service import create_contact, patch_contact
from app.utils.crm_validation import clean_optional_text, normalize_contact_email, normalize_lead_source
from app.utils.mongo_utils import object_id_or_not_found, public_document
from app.utils.phone_utils import normalize_indian_phone
from app.utils.time_utils import utc_now


MAX_IMPORT_FILE_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 10_000
MAX_IMPORT_COLUMNS = 100
IMPORT_ARTIFACT_HOURS = 24
IMPORT_SAMPLE_ROWS = 20

_TARGET_FIELDS = {
    "firstName", "lastName", "fullName", "phone", "alternatePhone", "email",
    "city", "state", "companyOrCollege", "instagramProfile", "facebookProfile",
    "linkedinProfile", "source", "courseInterest", "preferredMode",
    "targetExamYear", "notes",
}
_UPDATE_EMPTY_FIELDS = {
    "firstName", "lastName", "alternatePhone", "email", "city", "state",
    "companyOrCollege", "instagramProfile", "facebookProfile", "linkedinProfile", "notes",
}
_IMPORT_STATUSES = {
    item.value for item in LeadStatus if item not in {LeadStatus.ADMITTED, LeadStatus.DO_NOT_CONTACT}
}


def _aware(value: datetime) -> datetime:
    return value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value


def _safe_filename(filename: Optional[str]) -> str:
    value = (filename or "contacts").replace("\\", "/").split("/")[-1]
    value = re.sub(r"[^A-Za-z0-9._ -]+", "_", value).strip(" .")
    return (value or "contacts")[:180]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text[:2000]


def _validate_headers(headers: List[str]) -> List[str]:
    cleaned = [" ".join(value.strip().split()) for value in headers]
    if not cleaned or not any(cleaned):
        raise ValidationApiError("IMPORT_HEADERS_REQUIRED", "The file must contain a header row.")
    if len(cleaned) > MAX_IMPORT_COLUMNS:
        raise ValidationApiError("IMPORT_COLUMN_LIMIT_EXCEEDED", f"At most {MAX_IMPORT_COLUMNS} columns are supported.")
    if any(not value for value in cleaned):
        raise ValidationApiError("IMPORT_HEADER_BLANK", "Every import column must have a header.")
    if len({value.casefold() for value in cleaned}) != len(cleaned):
        raise ValidationApiError("IMPORT_HEADER_DUPLICATE", "File headers must be unique.")
    return cleaned


def _parse_csv(content: bytes) -> Tuple[List[str], List[Tuple[int, Dict[str, str]]]]:
    if content.startswith(b"PK"):
        raise ValidationApiError("IMPORT_CONTENT_MISMATCH", "The selected CSV file contains spreadsheet data.")
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValidationApiError("IMPORT_CSV_ENCODING_INVALID", "CSV files must use UTF-8 encoding.") from exc
    if "\x00" in decoded:
        raise ValidationApiError("IMPORT_CSV_MALFORMED", "The CSV file is malformed.")
    try:
        dialect = csv.Sniffer().sniff(decoded[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(decoded), dialect)
    try:
        headers = _validate_headers([_cell_text(value) for value in next(reader)])
    except StopIteration as exc:
        raise ValidationApiError("IMPORT_FILE_EMPTY", "The import file is empty.") from exc
    rows: List[Tuple[int, Dict[str, str]]] = []
    for row_number, row in enumerate(reader, start=2):
        if len(rows) >= MAX_IMPORT_ROWS:
            raise ValidationApiError("IMPORT_ROW_LIMIT_EXCEEDED", f"At most {MAX_IMPORT_ROWS} data rows are supported.")
        padded = list(row[: len(headers)]) + [""] * max(0, len(headers) - len(row))
        rows.append((row_number, dict(zip(headers, (_cell_text(value) for value in padded)))))
    return headers, rows


def _parse_xlsx(content: bytes) -> Tuple[List[str], List[Tuple[int, Dict[str, str]]]]:
    if not content.startswith(b"PK"):
        raise ValidationApiError("IMPORT_CONTENT_MISMATCH", "The selected XLSX file is not a valid spreadsheet.")
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
        sheet = workbook.active
        iterator = sheet.iter_rows()
        header_cells = next(iterator)
    except (StopIteration, KeyError, OSError, ValueError) as exc:
        raise ValidationApiError("IMPORT_XLSX_MALFORMED", "The XLSX file could not be read safely.") from exc
    headers = _validate_headers([_cell_text(cell.value) for cell in header_cells])
    rows: List[Tuple[int, Dict[str, str]]] = []
    for row_number, cells in enumerate(iterator, start=2):
        if len(rows) >= MAX_IMPORT_ROWS:
            workbook.close()
            raise ValidationApiError("IMPORT_ROW_LIMIT_EXCEEDED", f"At most {MAX_IMPORT_ROWS} data rows are supported.")
        selected = list(cells[: len(headers)])
        if any(getattr(cell, "data_type", None) == "f" for cell in selected):
            workbook.close()
            raise ValidationApiError(
                "IMPORT_FORMULA_UNSUPPORTED",
                f"Formula cells are not supported (row {row_number}). Save calculated values before import.",
            )
        values = [_cell_text(cell.value) for cell in selected]
        values += [""] * max(0, len(headers) - len(values))
        rows.append((row_number, dict(zip(headers, values))))
    workbook.close()
    return headers, rows


def analyze_import(
    *, filename: Optional[str], content_type: Optional[str], content: bytes, actor: Dict[str, Any], request_id: Optional[str]
) -> Dict[str, Any]:
    if not content:
        raise ValidationApiError("IMPORT_FILE_EMPTY", "Select a non-empty CSV or XLSX file.")
    if len(content) > MAX_IMPORT_FILE_BYTES:
        raise PayloadTooLargeError("IMPORT_FILE_TOO_LARGE", f"Import files must not exceed {MAX_IMPORT_FILE_BYTES // (1024 * 1024)} MiB.")
    safe_name = _safe_filename(filename)
    extension = PurePath(safe_name).suffix.casefold()
    if extension not in {".csv", ".xlsx"}:
        raise ValidationApiError("IMPORT_FILE_TYPE_UNSUPPORTED", "Only .csv and .xlsx files are supported.")
    media_type = (content_type or "").split(";", 1)[0].strip().casefold()
    allowed_media_types = {
        ".csv": {"", "text/csv", "application/csv", "text/plain", "application/vnd.ms-excel", "application/octet-stream"},
        ".xlsx": {"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/octet-stream"},
    }
    if media_type not in allowed_media_types[extension]:
        raise ValidationApiError("IMPORT_CONTENT_TYPE_INVALID", "The uploaded content type does not match a supported CSV or XLSX file.")
    headers, rows = _parse_csv(content) if extension == ".csv" else _parse_xlsx(content)
    if not rows:
        raise ValidationApiError("IMPORT_ROWS_REQUIRED", "The file must contain at least one data row.")
    now = utc_now()
    expires_at = now + timedelta(hours=IMPORT_ARTIFACT_HOURS)
    job = insert_import_job(
        {
            "entityType": "CONTACT_IMPORT",
            "status": "ANALYZED",
            "originalFilename": safe_name,
            "fileType": extension[1:].upper(),
            "contentType": (content_type or "application/octet-stream")[:100],
            "fileSize": len(content),
            "fileHash": hashlib.sha256(content).hexdigest(),
            "headers": headers,
            "sampleRows": [values for _, values in rows[:IMPORT_SAMPLE_ROWS]],
            "totalRows": len(rows),
            "artifactExpiresAt": expires_at,
            "createdBy": actor["_id"],
            "createdAt": now,
            "updatedAt": now,
        }
    )
    insert_import_rows(
        [
            {
                "importId": job["_id"],
                "rowNumber": row_number,
                "rawValues": values,
                "validationStatus": "UNREVIEWED",
                "expiresAt": expires_at,
                "createdAt": now,
            }
            for row_number, values in rows
        ]
    )
    write_audit_event(
        "CONTACT_IMPORT_ANALYZED",
        "SUCCEEDED",
        actor_user_id=actor["_id"],
        entity_type="CONTACT_IMPORT",
        entity_id=job["_id"],
        request_id=request_id,
        compact_metadata={"fileType": job["fileType"], "fileSize": len(content), "rowCount": len(rows)},
        operation_id=f"contact-import-analyze:{job['_id']}",
    )
    return job


def _get_job(import_id_value: Any, *, require_artifact: bool = False) -> Dict[str, Any]:
    import_id = object_id_or_not_found(import_id_value, "contact import")
    job = find_import_job(import_id)
    if not job:
        raise NotFoundError("CONTACT_IMPORT_NOT_FOUND", "The requested Contact import was not found.")
    if require_artifact and _aware(job["artifactExpiresAt"]) <= utc_now():
        update_import_job(import_id, {"status": "EXPIRED", "updatedAt": utc_now()})
        raise GoneError("CONTACT_IMPORT_EXPIRED", "The temporary import data has expired. Analyze the file again.")
    return job


def _mapped_value(raw: Dict[str, str], mapping: Dict[str, str], target: str) -> Optional[str]:
    header = mapping.get(target)
    return clean_optional_text(raw.get(header)) if header else None


def _normalize_mode(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    normalized = re.sub(r"[^A-Z0-9]+", "_", value.upper()).strip("_")
    aliases = {"BOTH": "HYBRID", "ONLINE_OFFLINE": "HYBRID", "OFFLINE_ONLINE": "HYBRID"}
    normalized = aliases.get(normalized, normalized)
    if normalized not in {item.value for item in PreferredMode}:
        raise ValidationApiError("PREFERRED_MODE_INVALID", "Preferred mode must be ONLINE, OFFLINE, or HYBRID.")
    return normalized


def _normalize_year(value: Optional[str], default: Optional[int]) -> Optional[int]:
    if value is None:
        return default
    try:
        number = float(value)
        if not number.is_integer():
            raise ValueError
        year = int(number)
    except (TypeError, ValueError) as exc:
        raise ValidationApiError("TARGET_YEAR_INVALID", "Target exam year must be a whole year.") from exc
    if not 2020 <= year <= 2100:
        raise ValidationApiError("TARGET_YEAR_INVALID", "Target exam year must be between 2020 and 2100.")
    return year


def _normalize_row(
    raw: Dict[str, str], mapping: Dict[str, str], defaults: Dict[str, Any], default_source: str
) -> Dict[str, Any]:
    if not any(clean_optional_text(value) for value in raw.values()):
        raise ValidationApiError("IMPORT_ROW_EMPTY", "The row is empty.")
    phone = _mapped_value(raw, mapping, "phone")
    normalized_phone = normalize_indian_phone(phone, "phone")
    first_name = _mapped_value(raw, mapping, "firstName")
    last_name = _mapped_value(raw, mapping, "lastName")
    full_name = _mapped_value(raw, mapping, "fullName")
    if full_name and not (first_name or last_name):
        parts = full_name.split(maxsplit=1)
        first_name = parts[0]
        last_name = parts[1] if len(parts) > 1 else None
    alternate_phone = _mapped_value(raw, mapping, "alternatePhone")
    if alternate_phone and normalize_indian_phone(alternate_phone, "alternatePhone") == normalized_phone:
        raise ValidationApiError("CONTACT_ALTERNATE_PHONE_DUPLICATE", "Alternate phone must differ from primary phone.")
    email = _mapped_value(raw, mapping, "email")
    normalize_contact_email(email)
    mapped_source = _mapped_value(raw, mapping, "source")
    source = normalize_lead_source(mapped_source or defaults.get("source"), default=default_source)
    mode = _normalize_mode(
        _mapped_value(raw, mapping, "preferredMode") or defaults.get("preferredMode")
    )
    target_year = _normalize_year(
        _mapped_value(raw, mapping, "targetExamYear"), defaults.get("targetExamYear")
    )
    status = defaults.get("status", LeadStatus.NEW.value)
    if status not in _IMPORT_STATUSES:
        raise ValidationApiError("IMPORT_LEAD_STATUS_INVALID", "The selected default Lead status is not available for import.")
    priority = defaults.get("priority", LeadPriority.MEDIUM.value)
    result = {
        "firstName": first_name,
        "lastName": last_name,
        "phone": phone,
        "normalizedPhone": normalized_phone,
        "alternatePhone": alternate_phone,
        "email": email,
        "city": _mapped_value(raw, mapping, "city"),
        "state": _mapped_value(raw, mapping, "state"),
        "companyOrCollege": _mapped_value(raw, mapping, "companyOrCollege"),
        "instagramProfile": _mapped_value(raw, mapping, "instagramProfile"),
        "facebookProfile": _mapped_value(raw, mapping, "facebookProfile"),
        "linkedinProfile": _mapped_value(raw, mapping, "linkedinProfile"),
        "source": source,
        "sourceDetails": clean_optional_text(defaults.get("sourceDetails")),
        "courseInterest": _mapped_value(raw, mapping, "courseInterest"),
        "preferredMode": mode,
        "targetExamYear": target_year,
        "notes": _mapped_value(raw, mapping, "notes"),
        "leadStatus": status,
        "leadPriority": priority,
        "assignedCounsellorId": defaults.get("assignedCounsellorId"),
    }
    return {key: value for key, value in result.items() if value is not None}


def _empty_field_updates(existing: Dict[str, Any], normalized: Dict[str, Any]) -> Dict[str, Any]:
    updates: Dict[str, Any] = {}
    for field in _UPDATE_EMPTY_FIELDS:
        value = normalized.get(field)
        if not value or existing.get(field) not in {None, ""}:
            continue
        if field in {"firstName", "lastName"} and existing.get("displayName"):
            continue
        updates[field] = value
    return updates


def preview_import(
    import_id_value: Any,
    payload: ContactImportPreviewModel,
    actor: Dict[str, Any],
    request_id: Optional[str],
    *, page: int, page_size: int,
) -> Dict[str, Any]:
    job = _get_job(import_id_value, require_artifact=True)
    if job.get("status") in {"COMMITTING", "COMPLETED"}:
        raise ConflictError("CONTACT_IMPORT_ALREADY_COMMITTED", "A committed import cannot be previewed again.")
    mapping = payload.mapping
    if set(mapping.values()) - set(job["headers"]):
        raise ValidationApiError("IMPORT_MAPPING_HEADER_INVALID", "A mapped column is not present in the uploaded file.")
    defaults = payload.defaults.model_dump(mode="json")
    if defaults.get("status") not in _IMPORT_STATUSES:
        raise ValidationApiError("IMPORT_LEAD_STATUS_INVALID", "The selected default Lead status is not available for import.")
    if defaults.get("assignedCounsellorId"):
        validate_counsellor(defaults["assignedCounsellorId"])
    default_source = "CSV_IMPORT" if job["fileType"] == "CSV" else "EXCEL_IMPORT"
    seen_phones = set()
    counts = {
        "totalRows": job["totalRows"], "validRows": 0, "existingDuplicates": 0,
        "withinFileDuplicates": 0, "updateEligibleRows": 0, "rejectedRows": 0,
    }
    for row in all_import_rows(job["_id"]):
        updates: Dict[str, Any]
        try:
            normalized = _normalize_row(row["rawValues"], mapping, defaults, default_source)
            phone = normalized["normalizedPhone"]
            if phone in seen_phones:
                counts["withinFileDuplicates"] += 1
                updates = {
                    "validationStatus": "DUPLICATE_IN_FILE",
                    "normalizedData": normalized,
                    "errors": [{"code": "DUPLICATE_IN_FILE", "message": "The normalized phone occurs more than once in this file."}],
                }
            else:
                seen_phones.add(phone)
                existing = find_contact_by_normalized_phone(phone)
                if existing:
                    counts["existingDuplicates"] += 1
                    empty_updates = _empty_field_updates(existing, normalized) if payload.duplicateMode == "UPDATE_EMPTY_FIELDS" else {}
                    status = "DUPLICATE_UPDATE" if empty_updates else "DUPLICATE_EXISTING"
                    if empty_updates:
                        counts["updateEligibleRows"] += 1
                    updates = {
                        "validationStatus": status,
                        "normalizedData": normalized,
                        "existingContactId": existing["_id"],
                        "updateFields": empty_updates,
                        "errors": [],
                    }
                else:
                    counts["validRows"] += 1
                    updates = {"validationStatus": "VALID", "normalizedData": normalized, "errors": []}
        except ApiError as exc:
            counts["rejectedRows"] += 1
            updates = {
                "validationStatus": "REJECTED",
                "errors": [{"code": exc.code, "message": exc.message, "fieldErrors": exc.field_errors}],
            }
        update_import_row(row["_id"], {**updates, "validatedAt": utc_now()})
    now = utc_now()
    job = update_import_job(
        job["_id"],
        {
            "status": "PREVIEWED", "mapping": mapping, "defaults": defaults,
            "duplicateMode": payload.duplicateMode, "counts": counts,
            "previewedAt": now, "updatedAt": now,
        },
    )
    write_audit_event(
        "CONTACT_IMPORT_PREVIEWED", "SUCCEEDED", actor_user_id=actor["_id"],
        entity_type="CONTACT_IMPORT", entity_id=job["_id"], request_id=request_id,
        compact_metadata=counts, operation_id=f"contact-import-preview:{job['_id']}:{uuid.uuid4()}",
    )
    return import_detail(job["_id"], page=page, page_size=page_size, require_artifact=True)


def _row_public(row: Dict[str, Any]) -> Dict[str, Any]:
    item = public_document(row)
    item.pop("expiresAt", None)
    item.pop("createdAt", None)
    return item


def import_detail(
    import_id_value: Any, *, page: int, page_size: int, require_artifact: bool = False
) -> Dict[str, Any]:
    job = _get_job(import_id_value, require_artifact=require_artifact)
    job_public = public_document(job)
    job_public.pop("fileHash", None)
    rows: List[Dict[str, Any]] = []
    pagination = pagination_metadata(page, page_size, 0)
    if _aware(job["artifactExpiresAt"]) > utc_now():
        documents, total = list_import_rows(job["_id"], page=page, page_size=page_size)
        rows = [_row_public(document) for document in documents]
        pagination = pagination_metadata(page, page_size, total)
    return {"import": job_public, "rows": rows, "pagination": pagination}


def commit_import(
    import_id_value: Any, actor: Dict[str, Any], request_id: Optional[str]
) -> Dict[str, Any]:
    job = _get_job(import_id_value, require_artifact=True)
    if job.get("status") == "COMPLETED":
        return job
    now = utc_now()
    claimed = claim_import_commit(job["_id"], now, now + timedelta(minutes=5))
    if not claimed:
        current = _get_job(job["_id"])
        if current.get("status") == "COMPLETED":
            return current
        raise ConflictError("CONTACT_IMPORT_COMMIT_IN_PROGRESS", "This import is already being committed. Check its status shortly.")
    preview_counts = claimed.get("counts", {})
    counts = {
        "totalRows": claimed["totalRows"],
        "validRows": preview_counts.get("validRows", 0),
        "importedContacts": 0, "importedLeads": 0,
        "existingDuplicates": preview_counts.get("existingDuplicates", 0),
        "updatedContacts": 0, "updatedEmptyFields": 0,
        "rejectedRows": preview_counts.get("rejectedRows", 0) + preview_counts.get("withinFileDuplicates", 0),
        "failedRows": 0,
    }
    for row in all_import_rows(claimed["_id"]):
        status = row.get("validationStatus")
        if status not in {"VALID", "DUPLICATE_UPDATE"}:
            continue
        try:
            if status == "DUPLICATE_UPDATE":
                existing = find_contact_by_id(row.get("existingContactId"))
                update_fields = row.get("updateFields") or {}
                if existing and update_fields:
                    patch_contact(
                        existing["_id"],
                        ContactPatchModel(version=int(existing.get("version", 1)), **update_fields),
                        actor,
                        request_id,
                    )
                    counts["updatedContacts"] += 1
                    counts["updatedEmptyFields"] += len(update_fields)
                update_import_row(row["_id"], {"validationStatus": "UPDATED", "committedAt": utc_now()})
                continue
            normalized = dict(row["normalizedData"])
            normalized.pop("normalizedPhone", None)
            created = create_contact(
                ContactCreateModel(**normalized),
                actor,
                request_id,
                operation_id=f"contact-import:{claimed['_id']}:{row['rowNumber']}",
                contact_metadata={"importJobId": claimed["_id"], "importRowNumber": row["rowNumber"]},
            )
            counts["importedContacts"] += 1
            counts["importedLeads"] += 1 if created.get("lead") else 0
            update_import_row(
                row["_id"],
                {
                    "validationStatus": "IMPORTED",
                    "contactId": created["contact"]["_id"],
                    "leadId": created["lead"]["_id"] if created.get("lead") else None,
                    "committedAt": utc_now(),
                },
            )
        except ConflictError as exc:
            if exc.code in {"CONTACT_PHONE_DUPLICATE", "ACTIVE_LEAD_DUPLICATE"}:
                counts["existingDuplicates"] += 1
                update_import_row(
                    row["_id"],
                    {"validationStatus": "DUPLICATE_RACE", "errors": [{"code": exc.code, "message": exc.message}]},
                )
            else:
                counts["failedRows"] += 1
                update_import_row(
                    row["_id"],
                    {"validationStatus": "FAILED", "errors": [{"code": exc.code, "message": exc.message}]},
                )
        except ApiError as exc:
            counts["failedRows"] += 1
            update_import_row(
                row["_id"],
                {"validationStatus": "FAILED", "errors": [{"code": exc.code, "message": exc.message}]},
            )
        except Exception:
            counts["failedRows"] += 1
            update_import_row(
                row["_id"],
                {"validationStatus": "FAILED", "errors": [{"code": "IMPORT_ROW_FAILED", "message": "The row could not be imported safely."}]},
            )
    completed_at = utc_now()
    result = update_import_job(
        claimed["_id"],
        {
            "status": "COMPLETED", "counts": counts, "completedAt": completed_at,
            "updatedAt": completed_at, "commitOperationId": f"contact-import-commit:{claimed['_id']}",
            "commitLeaseUntil": completed_at,
        },
    )
    write_audit_event(
        "CONTACT_IMPORT_COMMITTED", "SUCCEEDED", actor_user_id=actor["_id"],
        entity_type="CONTACT_IMPORT", entity_id=claimed["_id"], request_id=request_id,
        compact_metadata=counts, operation_id=f"contact-import-commit:{claimed['_id']}",
    )
    return result


def rejection_report(import_id_value: Any) -> Tuple[str, bytes]:
    job = _get_job(import_id_value, require_artifact=True)
    rows, _ = list_import_rows(
        job["_id"], page=1, page_size=MAX_IMPORT_ROWS,
        statuses=["REJECTED", "DUPLICATE_IN_FILE", "FAILED", "DUPLICATE_RACE"],
    )
    output = io.StringIO(newline="")
    headers = ["Original row number", *job["headers"], "Error code", "Error message"]
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        errors = row.get("errors") or [{"code": "IMPORT_ROW_REJECTED", "message": "Row rejected."}]
        raw_values = []
        for header in job["headers"]:
            value = str(row.get("rawValues", {}).get(header, ""))
            if value.startswith(("=", "+", "-", "@")):
                value = "'" + value
            raw_values.append(value)
        writer.writerow(
            [row["rowNumber"], *raw_values, "; ".join(item.get("code", "") for item in errors),
             "; ".join(item.get("message", "") for item in errors)]
        )
    filename = f"{PurePath(job['originalFilename']).stem}-rejections.csv"
    return _safe_filename(filename), output.getvalue().encode("utf-8-sig")
